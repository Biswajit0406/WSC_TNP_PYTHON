import pytest
from openpyxl import load_workbook
import os
import time
class utility:
    @staticmethod
    def excel_data(file_path,sheet_name):
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]

        data = []
        for row in sheet.iter_rows(min_row=2,values_only=True):
            data.append(row)
        return data






