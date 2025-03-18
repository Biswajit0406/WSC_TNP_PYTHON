import pytest
from openpyxl import load_workbook
import os
import time
class utility:
    @staticmethod
    def excel_data(file_path,sheet_name):
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]

        data = []
        for row in sheet.iter_rows(min_row=2,values_only=True):
            data.append(row)
        return data




    @staticmethod
    def take_screenshot(driver, name="Screenshot"):
        """Capture a screenshot and save it with a timestamp."""
        timestamp = time.strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"Screenshots/{name}_{timestamp}.png"

        # Ensure the screenshots directory exists
        if not os.path.exists("screenshots"):
            os.makedirs("Screenshots")

        driver.save_screenshot(filename)
        print(f"Screenshot saved at: {filename}")

